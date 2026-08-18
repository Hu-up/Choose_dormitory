from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def dorm_id(prefix: str, room_name: str) -> str:
    suffix = re.sub(r"[^0-9A-Za-z]+", "-", room_name).strip("-")
    return f"{prefix}-{suffix}"


def extract_rooms(workbook, sheet_name: str, gender: str, prefix: str) -> list[dict]:
    sheet = workbook[sheet_name]
    current_room = ""
    counts: dict[str, int] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        room_name = text(row[2])
        bed_name = text(row[3])
        if room_name:
            current_room = room_name
        if current_room and bed_name:
            counts[current_room] = counts.get(current_room, 0) + 1

    return [
        {
            "id": dorm_id(prefix, room_name),
            "name": room_name,
            "gender": gender,
            "capacity": capacity,
        }
        for room_name, capacity in counts.items()
    ]


def extract_students(workbook, sheet_name: str, gender: str) -> list[dict]:
    sheet = workbook[sheet_name]
    students = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        student_id = text(row[1])
        name = text(row[2])
        if student_id and name:
            students.append(
                {
                    "student_id": student_id,
                    "name": name,
                    "gender": gender,
                }
            )

    return students


def sql_value(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def write_js(output_path: Path, dorms: list[dict], students: list[dict]) -> None:
    content = (
        "window.DORM_DEFAULTS = "
        + json.dumps(dorms, ensure_ascii=False, indent=2)
        + ";\n\n"
        + "window.ALLOWED_STUDENTS = "
        + json.dumps(students, ensure_ascii=False, indent=2)
        + ";\n"
    )
    output_path.write_text(content, encoding="utf-8")


def write_sql(output_path: Path, dorms: list[dict], students: list[dict]) -> None:
    dorm_ids = ", ".join(sql_value(dorm["id"]) for dorm in dorms)
    student_ids = ", ".join(sql_value(student["student_id"]) for student in students)
    dorm_rows = ",\n".join(
        "  ({id}, {name}, {gender}, {capacity})".format(
            id=sql_value(dorm["id"]),
            name=sql_value(dorm["name"]),
            gender=sql_value(dorm["gender"]),
            capacity=dorm["capacity"],
        )
        for dorm in dorms
    )
    student_rows = ",\n".join(
        "  ({student_id}, {name}, {gender})".format(
            student_id=sql_value(student["student_id"]),
            name=sql_value(student["name"]),
            gender=sql_value(student["gender"]),
        )
        for student in students
    )

    content = f"""create extension if not exists pgcrypto;

create table if not exists public.dorms (
  id text primary key,
  name text not null unique,
  gender text not null check (gender in ('男', '女')),
  capacity integer not null check (capacity > 0),
  created_at timestamptz not null default now()
);

create table if not exists public.records (
  id uuid primary key default gen_random_uuid(),
  name text not null,
  gender text not null check (gender in ('男', '女')),
  student_id text not null,
  dorm_id text not null references public.dorms(id) on delete cascade,
  created_at timestamptz not null default now(),
  updated_at timestamptz not null default now()
);

create table if not exists public.allowed_students (
  student_id text primary key check (student_id ~ '^120242227[0-9]{{3}}$'),
  name text not null,
  gender text not null check (gender in ('男', '女')),
  created_at timestamptz not null default now(),
  updated_at timestamptz not null default now()
);

create table if not exists public.records_backup_before_excel_sync as
select *
from public.records
where false;

insert into public.records_backup_before_excel_sync
select r.*
from public.records r
where not exists (
  select 1
  from public.records_backup_before_excel_sync b
  where b.id = r.id
);

insert into public.dorms (id, name, gender, capacity)
values
{dorm_rows}
on conflict (id) do update
set name = excluded.name,
    gender = excluded.gender,
    capacity = excluded.capacity;

delete from public.records
where student_id not in ({student_ids});

delete from public.dorms
where id not in ({dorm_ids});

insert into public.allowed_students (student_id, name, gender)
values
{student_rows}
on conflict (student_id) do update
set name = excluded.name,
    gender = excluded.gender,
    updated_at = now();

delete from public.allowed_students
where student_id not in ({student_ids});

delete from public.records r
using public.allowed_students s
where lower(r.student_id) = lower(s.student_id)
  and (trim(r.name) <> trim(s.name) or r.gender <> s.gender);

drop index if exists public.records_student_unique;
drop index if exists public.records_student_id_unique;

create unique index if not exists records_student_id_unique
  on public.records (lower(student_id));

alter table public.records
  drop constraint if exists records_student_id_format;

alter table public.records
  add constraint records_student_id_format
  check (student_id ~ '^120242227[0-9]{{3}}$');

create or replace function public.touch_updated_at()
returns trigger
language plpgsql
set search_path = public
as $$
begin
  new.updated_at = now();
  return new;
end;
$$;

drop trigger if exists records_touch_updated_at on public.records;

create trigger records_touch_updated_at
before update on public.records
for each row
execute function public.touch_updated_at();

create or replace function public.choose_dorm(
  p_name text,
  p_gender text,
  p_student_id text,
  p_dorm_id text
)
returns jsonb
language plpgsql
security definer
set search_path = public
as $$
declare
  v_dorm public.dorms%rowtype;
  v_student public.allowed_students%rowtype;
  v_existing public.records%rowtype;
  v_occupied integer;
  v_status text;
begin
  if trim(p_student_id) !~ '^120242227[0-9]{{3}}$' then
    return jsonb_build_object('status', 'invalid_student_id');
  end if;

  select *
    into v_student
    from public.allowed_students
    where lower(student_id) = lower(trim(p_student_id));

  if not found then
    return jsonb_build_object('status', 'not_allowed');
  end if;

  if v_student.gender <> p_gender then
    return jsonb_build_object('status', 'student_gender_mismatch');
  end if;

  if trim(v_student.name) <> trim(p_name) then
    return jsonb_build_object('status', 'student_name_mismatch');
  end if;

  perform pg_advisory_xact_lock(hashtext(lower(trim(p_student_id))));

  select *
    into v_dorm
    from public.dorms
    where id = p_dorm_id
    for update;

  if not found then
    return jsonb_build_object('status', 'not_found');
  end if;

  if v_dorm.gender <> p_gender then
    return jsonb_build_object('status', 'gender_mismatch');
  end if;

  select *
    into v_existing
    from public.records
    where lower(student_id) = lower(trim(p_student_id))
    for update;

  select count(*)
    into v_occupied
    from public.records
    where dorm_id = p_dorm_id
      and (v_existing.id is null or id <> v_existing.id);

  if v_occupied >= v_dorm.capacity then
    return jsonb_build_object('status', 'full');
  end if;

  if v_existing.id is null then
    insert into public.records (name, gender, student_id, dorm_id)
    values (trim(p_name), p_gender, trim(p_student_id), p_dorm_id);
    v_status := 'created';
  else
    update public.records
      set name = trim(p_name),
          gender = p_gender,
          student_id = trim(p_student_id),
          dorm_id = p_dorm_id
      where id = v_existing.id;
    v_status := 'updated';
  end if;

  return jsonb_build_object('status', v_status);
end;
$$;

alter table public.dorms enable row level security;
alter table public.records enable row level security;
alter table public.allowed_students enable row level security;

drop policy if exists "public read dorms" on public.dorms;
drop policy if exists "public insert dorms" on public.dorms;
drop policy if exists "public update dorms" on public.dorms;
drop policy if exists "public delete dorms" on public.dorms;
drop policy if exists "public read records" on public.records;
drop policy if exists "public insert records" on public.records;
drop policy if exists "public update records" on public.records;
drop policy if exists "public delete records" on public.records;
drop policy if exists "public read allowed students" on public.allowed_students;

create policy "public read dorms" on public.dorms
  for select using (true);

create policy "public insert dorms" on public.dorms
  for insert with check (true);

create policy "public update dorms" on public.dorms
  for update using (true) with check (true);

create policy "public delete dorms" on public.dorms
  for delete using (true);

create policy "public read records" on public.records
  for select using (true);

create policy "public insert records" on public.records
  for insert with check (true);

create policy "public update records" on public.records
  for update using (true) with check (true);

create policy "public delete records" on public.records
  for delete using (true);

create policy "public read allowed students" on public.allowed_students
  for select using (false);

grant usage on schema public to anon;
grant select, insert, update, delete on public.dorms to anon;
grant select, insert, update, delete on public.records to anon;
grant execute on function public.choose_dorm(text, text, text, text) to anon;

select
  'excel data ready' as status,
  (select count(*) from public.dorms) as dorm_count,
  (select coalesce(sum(capacity), 0) from public.dorms) as bed_count,
  (select count(*) from public.allowed_students) as allowed_student_count,
  (select count(*) from public.records) as selected_record_count;
"""
    output_path.write_text(content, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("--write-js", type=Path)
    parser.add_argument("--write-sql", type=Path)
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)

    dorms = (
        extract_rooms(workbook, "床位男165", "男", "m")
        + extract_rooms(workbook, "床位女57", "女", "f")
    )
    students = (
        extract_students(workbook, "名单男165", "男")
        + extract_students(workbook, "名单女57", "女")
    )

    summary = {
        "dorm_count": len(dorms),
        "bed_count": sum(dorm["capacity"] for dorm in dorms),
        "male_dorm_count": sum(dorm["gender"] == "男" for dorm in dorms),
        "female_dorm_count": sum(dorm["gender"] == "女" for dorm in dorms),
        "student_count": len(students),
        "students_by_gender": Counter(student["gender"] for student in students),
        "duplicate_student_ids": [
            student_id
            for student_id, count in Counter(
                student["student_id"] for student in students
            ).items()
            if count > 1
        ],
        "first_dorms": dorms[:5],
        "last_dorms": dorms[-5:],
        "first_students": students[:3],
        "last_students": students[-3:],
    }

    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if args.write_js:
        write_js(args.write_js, dorms, students)
    if args.write_sql:
        write_sql(args.write_sql, dorms, students)


if __name__ == "__main__":
    main()
